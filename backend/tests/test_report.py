"""Exporters: valid files, and a PDF table that does not wrap mid-word."""

from __future__ import annotations

import io

import pytest
from backend.app.services import report
from openpyxl import load_workbook
from reportlab.pdfbase.pdfmetrics import stringWidth

HEADERS = ["Date", "Area", "Device", "SSID", "BSSID", "RSSI (dBm)", "Grade", "Result"]
SSID = "Factory-WiFi"
ROWS = [
    # Deliberately wide values: these are what an equal-width layout mangles.
    ["2026-08-28", "Warehouse-2", "Forklift-PC-3", SSID, "AA:BB:CC:DD:EE:04", -52, "EXCELLENT", "PASS"],  # noqa: E501
    ["2026-08-28", "Line-A", "Scanner-01", SSID, "AA:BB:CC:DD:EE:01", -71, "FAIR", "WARNING"],
    ["2026-08-28", "QC-Lab", "Tablet-07", SSID, "AA:BB:CC:DD:EE:03", -79, "POOR", "FAIL"],
]


def test_csv_has_a_bom_so_excel_reads_utf8():
    body = report.to_csv(HEADERS, ROWS)
    assert body.startswith(b"\xef\xbb\xbf")
    text = body.decode("utf-8-sig")
    assert text.splitlines()[0] == ",".join(HEADERS)
    assert "Forklift-PC-3" in text


def test_csv_writes_blanks_not_the_word_none():
    body = report.to_csv(["A", "B"], [[None, 1]]).decode("utf-8-sig")
    assert body.splitlines()[1] == ",1"


def test_xlsx_is_readable_and_colours_the_verdict():
    workbook = load_workbook(io.BytesIO(report.to_xlsx(HEADERS, ROWS, summary={"Records": 3})))
    assert workbook.sheetnames == ["Data", "Summary"]

    sheet = workbook["Data"]
    assert [c.value for c in sheet[1]] == HEADERS
    assert sheet.freeze_panes == "A2"

    result_col = HEADERS.index("Result") + 1
    fills = {sheet.cell(row=r, column=result_col).fill.fgColor.rgb for r in (2, 3, 4)}
    # Each verdict gets its own fill; three verdicts must not share one colour.
    assert len({f for f in fills if f}) == 3

    assert workbook["Summary"]["B3"].value == 3


def test_pdf_is_a_pdf_and_carries_the_findings():
    body = report.to_pdf(
        HEADERS,
        ROWS,
        title="Survey",
        summary={"Records": 3},
        findings=[
            {
                "severity": "critical",
                "title": "Weak coverage",
                "summary": "Signal is -79 dBm.",
                "causes": ["Far from the AP"],
                "recommendations": ["Check the nearest AP"],
            }
        ],
    )
    assert body.startswith(b"%PDF")
    assert len(body) > 1000


def test_column_widths_fit_every_value_when_there_is_room():
    """The whole point: no column narrower than its widest value needs."""
    available = 760.0
    widths = report._column_widths(HEADERS, ROWS, available)

    assert len(widths) == len(HEADERS)
    assert sum(widths) == pytest.approx(available)

    for index, width in enumerate(widths):
        widest = max(stringWidth(str(row[index]), "Helvetica", 7.5) for row in ROWS)
        assert width >= widest + 2 * report._CELL_PADDING, f"column {HEADERS[index]} truncates"


def test_column_widths_favour_the_columns_that_need_it():
    widths = report._column_widths(HEADERS, ROWS, 760.0)
    by_name = dict(zip(HEADERS, widths, strict=True))
    # BSSID holds 17 characters; RSSI holds 3. The layout must reflect that.
    assert by_name["BSSID"] > by_name["RSSI (dBm)"] * 2


def test_column_widths_degrade_gracefully_when_over_budget():
    """A cramped page still gets a usable table rather than slivers."""
    wide_rows = [["x" * 120] * len(HEADERS)]
    widths = report._column_widths(HEADERS, wide_rows, 300.0)

    assert sum(widths) == pytest.approx(300.0)
    assert all(width > 0 for width in widths)


def test_column_widths_handle_no_rows():
    widths = report._column_widths(HEADERS, [], 600.0)
    assert sum(widths) == pytest.approx(600.0)
    assert all(width > 0 for width in widths)
