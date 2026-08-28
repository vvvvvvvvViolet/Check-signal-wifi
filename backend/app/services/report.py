"""Export survey data as CSV, Excel or PDF.

The three formats serve different readers, so they are not the same table three
times: CSV is the raw dump for a spreadsheet pivot, Excel adds grade colouring
and a summary sheet, and the PDF is the one-page thing a manager signs.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable, Sequence
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .quality import GRADE_COLOR

RESULT_FILL = {
    "PASS": "C6EFCE",
    "WARNING": "FFEB9C",
    "FAIL": "FFC7CE",
}
RESULT_FONT = {
    "PASS": "006100",
    "WARNING": "9C6500",
    "FAIL": "9C0006",
}


def _fmt(value) -> str:
    if value is None:
        return "-"
    if isinstance(value, datetime):
        return value.astimezone(UTC).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float):
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


# ------------------------------------------------------------------- CSV
def to_csv(headers: Sequence[str], rows: Iterable[Sequence]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(headers)
    for row in rows:
        writer.writerow(["" if v is None else v for v in row])
    # BOM so Excel opens UTF-8 (Thai area names) correctly on a double click.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


# ----------------------------------------------------------------- Excel
def to_xlsx(
    headers: Sequence[str],
    rows: Iterable[Sequence],
    *,
    title: str = "WiFi Survey",
    summary: dict | None = None,
    result_column: str | None = "Result",
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, name in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    result_idx = headers.index(result_column) + 1 if result_column in headers else None
    widths = [len(str(h)) for h in headers]

    materialised = [list(row) for row in rows]
    for r, row in enumerate(materialised, start=2):
        for c, value in enumerate(row, start=1):
            cell = sheet.cell(row=r, column=c, value=value)
            widths[c - 1] = max(widths[c - 1], len(_fmt(value)))
        if result_idx is not None:
            verdict = str(row[result_idx - 1] or "").upper()
            if verdict in RESULT_FILL:
                target = sheet.cell(row=r, column=result_idx)
                target.fill = PatternFill("solid", fgColor=RESULT_FILL[verdict])
                target.font = Font(bold=True, color=RESULT_FONT[verdict])

    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = min(40, max(10, width + 3))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{len(materialised) + 1}"
        if materialised
        else f"A1:{get_column_letter(len(headers))}1"
    )

    meta = workbook.create_sheet("Summary")
    meta["A1"] = title
    meta["A1"].font = Font(bold=True, size=14)
    meta["A2"] = "Generated"
    meta["B2"] = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")
    meta["A3"] = "Records"
    meta["B3"] = len(materialised)
    row = 5
    for key, value in (summary or {}).items():
        meta.cell(row=row, column=1, value=str(key)).font = Font(bold=True)
        meta.cell(
            row=row,
            column=2,
            value=value if isinstance(value, int | float) else _fmt(value),
        )
        row += 1
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 32

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


# ------------------------------------------------------------------- PDF
def to_pdf(
    headers: Sequence[str],
    rows: Iterable[Sequence],
    *,
    title: str = "WiFi Survey Report",
    subtitle: str | None = None,
    summary: dict | None = None,
    findings: list[dict] | None = None,
) -> bytes:
    stream = io.BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=18, spaceAfter=4)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.5, leading=9)

    story: list = [Paragraph(title, h1)]
    if subtitle:
        story.append(Paragraph(subtitle, sub))
    story.append(
        Paragraph(
            f"Generated {datetime.now(UTC).strftime('%Y-%m-%d %H:%M:%S UTC')}", sub
        )
    )
    story.append(Spacer(1, 6 * mm))

    if summary:
        summary_rows = [
            [Paragraph(f"<b>{k}</b>", cell), Paragraph(_fmt(v), cell)]
            for k, v in summary.items()
        ]
        table = Table(summary_rows, colWidths=[55 * mm, 75 * mm], hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d5dae1")),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5f9")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.extend([table, Spacer(1, 6 * mm)])

    materialised = [list(r) for r in rows]
    data = [[Paragraph(f"<b>{h}</b>", cell) for h in headers]]
    data += [[Paragraph(_fmt(v), cell) for v in row] for row in materialised]

    available = doc.width
    table = Table(data, repeatRows=1, colWidths=[available / len(headers)] * len(headers))
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d5dae1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9fb")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    # Tint the verdict column so a reviewer can scan the failures at a glance.
    if "Result" in headers:
        idx = headers.index("Result")
        for r, row in enumerate(materialised, start=1):
            verdict = str(row[idx] or "").upper()
            if verdict in RESULT_FILL:
                style.append(
                    ("BACKGROUND", (idx, r), (idx, r), colors.HexColor("#" + RESULT_FILL[verdict]))
                )
    table.setStyle(TableStyle(style))
    story.append(table)

    if findings:
        story.append(PageBreak())
        story.append(Paragraph("Diagnosis", styles["Heading2"]))
        for finding in findings:
            colour = {"critical": "#b91c1c", "warning": "#b45309"}.get(
                finding.get("severity", ""), "#334155"
            )
            story.append(
                Paragraph(
                    f'<font color="{colour}"><b>[{finding.get("severity", "info").upper()}] '
                    f'{finding.get("title", "")}</b></font>',
                    styles["Normal"],
                )
            )
            story.append(Paragraph(finding.get("summary", ""), cell))
            for label, key in (
                ("Possible problem", "causes"),
                ("Recommendation", "recommendations"),
            ):
                items = finding.get(key) or []
                if items:
                    story.append(Paragraph(f"<b>{label}:</b>", cell))
                    for item in items:
                        story.append(Paragraph(f"&bull; {item}", cell))
            story.append(Spacer(1, 4 * mm))

    doc.build(story)
    return stream.getvalue()


def grade_hex(grade: str | None) -> str:
    return GRADE_COLOR.get((grade or "").upper(), "#94a3b8")
